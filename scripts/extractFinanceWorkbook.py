#!/usr/bin/env python3
"""Extract the governed Finance workbooks as JSON without changing the sources."""

from __future__ import annotations

import argparse
import json
import math
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def normalized_text(value):
    if value is None:
        return None
    if isinstance(value, str):
        return unicodedata.normalize("NFC", value.strip())
    return value


def json_value(value):
    value = normalized_text(value)
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def rows_from_sheet(path: Path, sheet_name: str, id_header: str):
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_vba=path.suffix.lower() == ".xlsm",
    )
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    header_index = next(
        index
        for index, row in enumerate(values)
        if id_header in {normalized_text(cell) for cell in row if cell is not None}
    )
    headers = [
        unicodedata.normalize("NFC", cell) if isinstance(cell, str) else cell
        for cell in values[header_index]
    ]
    records = []
    source_rows = []
    for row_index, row in enumerate(values[header_index + 1 :], header_index + 2):
        record = {
            header: json_value(row[column_index])
            for column_index, header in enumerate(headers)
            if header and column_index < len(row)
        }
        identifier = record.get(id_header)
        if not identifier or not str(identifier).startswith(("REC-", "DEP-")):
            continue
        record["SOURCE_ROW"] = row_index
        records.append(record)
        source_rows.append(row_index)
    return records, source_rows


def department_for_bu(value):
    normalized = str(value or "").upper().replace("_", "")
    return {
        "ADMINORG": "Administration",
        "IMMO": "Finances",
        "SOCIAL": "Finances",
        "IMPORTEXPORT": "Commercial & CRM",
    }.get(normalized, str(value or ""))


def enrich_income(records):
    for record in records:
        record["DEPARTEMENT"] = record.get("DEPARTEMENT") or department_for_bu(record.get("BU"))
        record["SOURCE_FILE"] = "RECETTES.xlsm"
        record["SOURCE_SHEET"] = "BDD_Recettes"


def enrich_expenses(records):
    for record in records:
        record["DEPARTEMENT"] = record.get("DEPARTEMENT") or department_for_bu(record.get("BU"))
        if not record.get("TEAM"):
            country = str(record.get("PAYS") or "").upper()
            record["TEAM"] = "Team_ZH" if country == "CH" else "Team_SN" if country == "SN" else None
        record["SOURCE_FILE"] = "BDD_DEPENSES.xlsx"
        record["SOURCE_SHEET"] = "BDD_Depense"


def total(records, field):
    return round(sum(float(record.get(field) or 0) for record in records), 6)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--income", required=True, type=Path)
    parser.add_argument("--expenses", required=True, type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    income, _ = rows_from_sheet(args.income, "BDD_Recettes", "ID_RECETTE")
    expenses, _ = rows_from_sheet(args.expenses, "BDD_Depense", "Nr REF")
    enrich_income(income)
    enrich_expenses(expenses)

    social = [
        record
        for record in income
        if str(record.get("NATURE_RECETTE") or "").upper()
        in {"AIDE SOCIALE MÉNAGE", "AIDE SOCIALE MENAGE", "AIDE SOCIALE"}
    ]
    anomalies = {
        "income_missing_positive_chf": [
            record["ID_RECETTE"]
            for record in income
            if float(record.get("MONTANT_CFA") or 0) > 0 and float(record.get("MONTANT_CHF") or 0) <= 0
        ],
        "income_missing_fx": [
            record["ID_RECETTE"]
            for record in income
            if float(record.get("TAUX_FX_APPLIQUE") or 0) <= 0
        ],
    }
    payload = {
        "metadata": {
            "income_source": args.income.name,
            "expenses_source": args.expenses.name,
            "income_count": len(income),
            "expense_count": len(expenses),
            "income_total_chf": total(income, "MONTANT_CHF"),
            "income_total_cfa": total(income, "MONTANT_CFA"),
            "expense_total_chf": total(expenses, "CHF"),
            "expense_total_cfa": total(expenses, "CFA"),
            "social_count": len(social),
            "social_total_chf": total(social, "MONTANT_CHF"),
            "social_total_cfa": total(social, "MONTANT_CFA"),
            "anomalies": anomalies,
        },
        "income": income,
        "expenses": expenses,
    }
    if args.output:
        args.output.write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
    else:
        json.dump(payload, sys.stdout, ensure_ascii=False, separators=(",", ":"))


if __name__ == "__main__":
    main()
